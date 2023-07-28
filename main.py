from tkinter import *
from tkinter import filedialog as fd
import psycopg2
import openpyxl
import tkinter.messagebox as mb
import xml.etree.ElementTree as ET
import json
import io
from dicttoxml import dicttoxml
from xml.dom.minidom import parseString
my_database="diplom_1"
my_user="postgres"
my_password="kmike"
my_host="localhost"
my_port="5432"
n = 0
rows = []

def quest_def(quest_text):
    global my_database, my_user, my_password, my_host, my_port, n
    conn = psycopg2.connect(database=my_database, user=my_user, password=my_password, host=my_host, port=my_port)
    cur = conn.cursor()
    cur.execute(quest_text)
    arr = cur.fetchall()
    conn.close()
    return arr

class Form:
    def __init__(self, form_name, geom):
        self.root = Tk()
        self.root.title(form_name)
        self.root.geometry(geom)
        self.width = 50
class Form2:
    def __init__(self, form_name, geom):
        self.root = Tk()
        form_name1 = form_name[0:8]
        form_name2 = form_name[8:-0]
        self.root.title(form_name1 + form_name2)
        self.root.geometry(geom)
        self.width = 50
class Entry_1:
    def __init__(self, form, label, data, type):
        if type == True:
            self.lab_1 = Label(form.root, text=label)
        else:
            self.lab_1 = Label(form.root, text=label[2:-3])
        self.ent_1 = Entry(form.root, width=form.width + 16)
        self.lab_1.pack()
        self.ent_1.pack()
        self.ent_1.insert(0, str(data))
    def get_name(self):
        return self.lab_1.cget()
    def set_1(self, text):
        self.ent_1.delete(0, END)
        self.ent_1.insert(0, text)
    def get_1(self):
        return self.ent_1.get()
class Button_1:
    def __init__(self, form, text_b, width_b, height_b, command_b):
        self.b1 = Button(form.root, text=text_b, width=width_b, height=height_b)
        self.b1.config(command=command_b)
        self.b1.pack()

    def get_1(self):
        return self.b1.cget("text")
class Button_2:
    def __init__(self, form, text_b, width_b, height_b, command_b):
        self.b1 = Button(form.root, text=text_b[2:-3], width=width_b, height=height_b)
        self.b1.config(command=command_b)
        self.b1.pack()

class Scene_1:
    def __init__(self):
        self.form_login = Form('авторизация', '450x500+400+50')
        self.qstn_username = Entry_1(self.form_login, "user name", '',True)
        self.qstn_password = Entry_1(self.form_login, "password", '',True)
        self.but_1 = Button_1(self.form_login, "войти", 50, 1,
                              lambda: self.login(str(self.qstn_username.get_1()).strip(),
                                                 str(self.qstn_password.get_1()).strip()))
        self.form_login.root.mainloop()
    def login(self, name, password):
        quest = "select user_name, password, type_profile from login where user_name = '{0}' and password = '{1}';".format(
            name, password)
        quest_response = str((quest_def(quest))[0])
        if quest_response == "[]":
            self.show_error("вы ввели неправилный пароль или логин")
        elif (quest_response.split("\'"))[5] == "admin":
            get_access = ['orders', 'employees', 'customers', 'position', 'repair_part', 'repair_parts_to_order', 'login', 'repair_part_seller']
        elif (quest_response.split("\'"))[5] == "director":
            get_access = ['orders', 'employees', 'customers', 'position', 'repair_part', 'repair_parts_to_order', 'repair_part_seller']
        elif (quest_response.split("\'"))[5] == "personnel":
            get_access = ['employees', 'position']
        elif (quest_response.split("\'"))[5] == "manager":
            get_access = ['orders', 'customers', 'repair_part_seller', 'repair_part', 'repair_parts_to_order']
        elif (quest_response.split("\'"))[5] == "engineer":
            get_access = ['orders', 'repair_part', 'repair_part_to_order']
        elif (quest_response.split("\'"))[5] == "manager_purchasing":
            get_access = ['repair_part_seller', 'repair_part', 'repair_parts_to_order']
        elif (quest_response.split("\'"))[5] == "manager_client":
            get_access = ['orders', 'customers']
        else:
            self.show_error("несуществующий тип пользователя")
        self.call_menu(get_access)
    def call_menu(self, access):
        self.form_login.root.destroy()
        self.form_menu_access = Form("меню выбора меню таблиц", '500x700+400+0')
        quest = "select table_name from information_schema.tables where table_schema='public';"
        table_arr = quest_def(quest)
        for element in access:
            if element == "orders":
                but_1 = Button_1(self.form_menu_access, element, 50, 1, lambda: self.table_menu_1(table_arr[0]))
            if element == "employees":
                but_2 = Button_1(self.form_menu_access, element, 50, 1, lambda: self.table_menu_2(table_arr[1]))
            if element == "customers":
                but_3 = Button_1(self.form_menu_access, element, 50, 1, lambda: self.table_menu_3(table_arr[2]))
            if element == "position":
                but_4 = Button_1(self.form_menu_access, element, 50, 1, lambda: self.table_menu_4(table_arr[3]))
            if element == "repair_part":
                but_5 = Button_1(self.form_menu_access, element, 50, 1, lambda: self.table_menu_5(table_arr[4]))
            if element == "repair_parts_to_order":
                but_6 = Button_1(self.form_menu_access, element, 50, 1, lambda: self.table_menu_6(table_arr[5]))
            if element == "login":
                but_7 = Button_1(self.form_menu_access, element, 50, 1, lambda: self.table_menu_7(table_arr[6]))
            if element == "repair_part_seller":
                but_8 = Button_1(self.form_menu_access, element, 50, 1, lambda: self.table_menu_8(table_arr[7]))
        self.form_menu_access.root.mainloop()

    def connect(self, table_arr):
        global rows
        global my_database, my_user, my_password, my_host, my_port
        conn = psycopg2.connect(database=my_database, user=my_user, password=my_password, host=my_host, port=my_port)
        cur = conn.cursor()
        quest1 = "select * from {0}".format(table_arr)
        cur.execute(quest1)
        rows = cur.fetchall()
        conn.close()

    def quest_def(self,quest_text):
        global my_database, my_user, my_password, my_host, my_port, n
        conn = psycopg2.connect(database=my_database, user=my_user, password=my_password, host=my_host, port=my_port)
        cur = conn.cursor()
        cur.execute(quest_text)
        arr = cur.fetchall()
        conn.close()
        return arr
    def table_menu_1(self,table_arr):
        global rows, n
        n = 0
        quest = "SELECT column_name FROM information_schema.columns WHERE table_name = '{0}' order by ordinal_position;".format(table_arr[0])
        name_column = self.quest_def(quest)
        quest = "select * from {0}".format(table_arr[0])
        rows = self.quest_def(quest)
        n = 0
        self.form2 = Form('таблица {0}'.format(table_arr[0]),'500x700+400+0')
        self.qstn_1 = Entry_1(self.form2, '{0}'.format(name_column[0]), rows[n][0], False)
        if len(name_column) > 1:
            self.qstn_2 = Entry_1(self.form2, '{0}'.format(name_column[1]), rows[n][1], False)
            if len(name_column) > 2:
                self.qstn_3 = Entry_1(self.form2, '{0}'.format(name_column[2]), rows[n][2], False)
                if len(name_column) > 3:
                    self.qstn_4 = Entry_1(self.form2, '{0}'.format(name_column[3]), rows[n][3], False)
                    if len(name_column) > 4:
                        self.qstn_5 = Entry_1(self.form2, '{0}'.format(name_column[4]), rows[n][4], False)
                        if len(name_column) > 5:
                            self.qstn_6 = Entry_1(self.form2, '{0}'.format(name_column[5]), rows[n][5], False)
                            if len(name_column) > 6:
                                self.qstn_7 = Entry_1(self.form2, '{0}'.format(name_column[6]), rows[n][6], False)
                                if len(name_column) > 7:
                                    self.qstn_8 = Entry_1(self.form2, '{0}'.format(name_column[7]), rows[n][7], False)
                                    if len(name_column) > 8:
                                        self.qstn_9 = Entry_1(self.form2, '{0}'.format(name_column[8]), rows[n][8], False)
                                        if len(name_column) > 9:
                                            self.qstn_10 = Entry_1(self.form2, '{0}'.format(name_column[9]), rows[n][9], False)
                                            if len(name_column) > 10:
                                                self.qstn_11 = Entry_1(self.form2, '{0}'.format(name_column[10]), rows[n][10], False)
                                                if len(name_column) > 11:
                                                    self.qstn_12 = Entry_1(self.form2, '{0}'.format(name_column[11]),rows[n][11], False)
        self.but_1 = Button_1(self.form2, "Следующая запись", 50, 1, lambda: self.next_line(table_arr[0]))
        self.but_2 = Button_1(self.form2, "Предыдущяя запись", 50, 1, lambda: self.previous_line(table_arr[0]))
        self.but_9 = Button_1(self.form2, "открыть по id", 50, 1, lambda: self.open_id(table_arr, name_column[0], rows))
        self.but_3 = Button_1(self.form2, "добавить запись", 50, 1, lambda: self.save_data(table_arr[0], rows, name_column[0]))
        self.but_4 = Button_1(self.form2, "Удалить запись", 50, 1, lambda: self.del_data(table_arr[0], name_column[0]))
        self.but_5 = Button_1(self.form2, "сохранить в файл", 50, 1, lambda: self.extract_text(rows))
        self.but_6 = Button_1(self.form2, "сохранить изменения", 50, 1, lambda: self.save_changes_data(table_arr[0], name_column, rows))
        self.but_7 = Button_1(self.form2, "сохранить в таблицу", 50, 1, lambda: self.save_table(rows, table_arr[0]))
        self.but_8 = Button_1(self.form2, "загрузить из таблицу", 50, 1, lambda: self.load_table(rows, table_arr[0]))
    def table_menu_2(self,table_arr):
        global rows, n
        n = 0
        quest = "SELECT column_name FROM information_schema.columns WHERE table_name = '{0}' order by ordinal_position;".format(table_arr[0])
        name_column = self.quest_def(quest)
        quest = "select * from {0}".format(table_arr[0])
        rows = self.quest_def(quest)
        self.form2 = Form('таблица {0}'.format(table_arr[0]),'500x600+400+200')
        self.qstn_1 = Entry_1(self.form2, '{0}'.format(name_column[0]), rows[0][0], False)
        if len(name_column) > 1:
            self.qstn_2 = Entry_1(self.form2, '{0}'.format(name_column[1]), rows[0][1], False)
            if len(name_column) > 2:
                self.qstn_3 = Entry_1(self.form2, '{0}'.format(name_column[2]), rows[0][2], False)
                if len(name_column) > 3:
                    self.qstn_4 = Entry_1(self.form2, '{0}'.format(name_column[3]), rows[0][3], False)
                    if len(name_column) > 4:
                        self.qstn_5 = Entry_1(self.form2, '{0}'.format(name_column[4]), rows[0][4], False)
                        if len(name_column) > 5:
                            self.qstn_6 = Entry_1(self.form2, '{0}'.format(name_column[5]), rows[0][5], False)
                            if len(name_column) > 6:
                                self.qstn_7 = Entry_1(self.form2, '{0}'.format(name_column[6]), rows[0][6], False)
                                if len(name_column) > 7:
                                    self.qstn_8 = Entry_1(self.form2, '{0}'.format(name_column[7]), rows[0][7], False)
                                    if len(name_column) > 8:
                                        self.qstn_9 = Entry_1(self.form2, '{0}'.format(name_column[8]), rows[0][8], False)
                                        if len(name_column) > 9:
                                            self.qstn_10 = Entry_1(self.form2, '{0}'.format(name_column[9]), rows[0][9], False)
                                            if len(name_column) > 10:
                                                self.qstn_11 = Entry_1(self.form2, '{0}'.format(name_column[10]), rows[0][10], False)
        self.but_1 = Button_1(self.form2, "Следующая запись", 50, 1, lambda: self.next_line(table_arr[0]))
        self.but_2 = Button_1(self.form2, "Предыдущяя запись", 50, 1, lambda: self.previous_line(table_arr[0]))
        self.but_3 = Button_1(self.form2, "добавить запись", 50, 1, lambda: self.save_data(table_arr[0], rows))
        self.but_4 = Button_1(self.form2, "Удалить запись", 50, 1,
                              lambda: self.del_data(table_arr[0], name_column[0]))
        self.but_5 = Button_1(self.form2, "сохранить в файл", 50, 1, lambda: self.extract_text(rows))
        self.but_6 = Button_1(self.form2, "сохранить изменения", 50, 1,
                              lambda: self.save_changes_data(table_arr[0], name_column, rows))
        self.but_7 = Button_1(self.form2, "сохранить в таблицу", 50, 1, lambda: self.save_table(rows, table_arr[0]))
        self.but_8 = Button_1(self.form2, "загрузить из таблицу", 50, 1, lambda: self.load_table(rows, table_arr[0]))
    def table_menu_3(self,table_arr):
        global rows, n
        n = 0
        quest = "SELECT column_name FROM information_schema.columns WHERE table_name = '{0}' order by ordinal_position;".format(table_arr[0])
        name_column = self.quest_def(quest)
        quest = "select * from {0}".format(table_arr[0])
        rows = self.quest_def(quest)
        self.form2 = Form('таблица {0}'.format(table_arr[0]),'500x600+400+200')
        self.qstn_1 = Entry_1(self.form2, '{0}'.format(name_column[0]), rows[0][0], False)
        if len(name_column) > 1:
            self.qstn_2 = Entry_1(self.form2, '{0}'.format(name_column[1]), rows[0][1], False)
            if len(name_column) > 2:
                self.qstn_3 = Entry_1(self.form2, '{0}'.format(name_column[2]), rows[0][2], False)
                if len(name_column) > 3:
                    self.qstn_4 = Entry_1(self.form2, '{0}'.format(name_column[3]), rows[0][3], False)
                    if len(name_column) > 4:
                        self.qstn_5 = Entry_1(self.form2, '{0}'.format(name_column[4]), rows[0][4], False)
                        if len(name_column) > 5:
                            self.qstn_6 = Entry_1(self.form2, '{0}'.format(name_column[5]), rows[0][5], False)
                            if len(name_column) > 6:
                                self.qstn_7 = Entry_1(self.form2, '{0}'.format(name_column[6]), rows[0][6], False)
                                if len(name_column) > 7:
                                    self.qstn_8 = Entry_1(self.form2, '{0}'.format(name_column[7]), rows[0][7], False)
                                    if len(name_column) > 8:
                                        self.qstn_9 = Entry_1(self.form2, '{0}'.format(name_column[8]), rows[0][8], False)
                                        if len(name_column) > 9:
                                            self.qstn_10 = Entry_1(self.form2, '{0}'.format(name_column[9]), rows[0][9], False)
                                            if len(name_column) > 10:
                                                self.qstn_11 = Entry_1(self.form2, '{0}'.format(name_column[10]), rows[0][10], False)
        self.but_1 = Button_1(self.form2, "Следующая запись", 50, 1, lambda: self.next_line(table_arr[0]))
        self.but_2 = Button_1(self.form2, "Предыдущяя запись", 50, 1, lambda: self.previous_line(table_arr[0]))
        self.but_3 = Button_1(self.form2, "добавить запись", 50, 1, lambda: self.save_data(table_arr[0], rows))
        self.but_4 = Button_1(self.form2, "Удалить запись", 50, 1,
                              lambda: self.del_data(table_arr[0], name_column[0], rows))
        self.but_5 = Button_1(self.form2, "сохранить в файл", 50, 1, lambda: self.extract_text(rows))
        self.but_6 = Button_1(self.form2, "сохранить изменения", 50, 1,
                              lambda: self.save_changes_data(table_arr[0], name_column, rows))
        self.but_7 = Button_1(self.form2, "сохранить в таблицу", 50, 1, lambda: self.save_table(rows, table_arr[0]))
        self.but_8 = Button_1(self.form2, "загрузить из таблицу", 50, 1, lambda: self.load_table(rows, table_arr[0]))
    def table_menu_4(self,table_arr):
        global rows, n
        n = 0
        quest = "SELECT column_name FROM information_schema.columns WHERE table_name = '{0}' order by ordinal_position;".format(table_arr[0])
        name_column = self.quest_def(quest)
        quest = "select * from {0}".format(table_arr[0])
        rows = self.quest_def(quest)
        self.form2 = Form('таблица {0}'.format(table_arr[0]),'500x600+400+200')
        self.qstn_1 = Entry_1(self.form2, '{0}'.format(name_column[0]), rows[0][0], False)
        if len(name_column) > 1:
            self.qstn_2 = Entry_1(self.form2, '{0}'.format(name_column[1]), rows[0][1], False)
            if len(name_column) > 2:
                self.qstn_3 = Entry_1(self.form2, '{0}'.format(name_column[2]), rows[0][2], False)
                if len(name_column) > 3:
                    self.qstn_4 = Entry_1(self.form2, '{0}'.format(name_column[3]), rows[0][3], False)
                    if len(name_column) > 4:
                        self.qstn_5 = Entry_1(self.form2, '{0}'.format(name_column[4]), rows[0][4], False)
                        if len(name_column) > 5:
                            self.qstn_6 = Entry_1(self.form2, '{0}'.format(name_column[5]), rows[0][5], False)
                            if len(name_column) > 6:
                                self.qstn_7 = Entry_1(self.form2, '{0}'.format(name_column[6]), rows[0][6], False)
                                if len(name_column) > 7:
                                    self.qstn_8 = Entry_1(self.form2, '{0}'.format(name_column[7]), rows[0][7], False)
                                    if len(name_column) > 8:
                                        self.qstn_9 = Entry_1(self.form2, '{0}'.format(name_column[8]), rows[0][8], False)
                                        if len(name_column) > 9:
                                            self.qstn_10 = Entry_1(self.form2, '{0}'.format(name_column[9]), rows[0][9], False)
                                            if len(name_column) > 10:
                                                self.qstn_11 = Entry_1(self.form2, '{0}'.format(name_column[10]), rows[0][10], False)
        self.but_1 = Button_1(self.form2, "Следующая запись", 50, 1, lambda: self.next_line(table_arr[0]))
        self.but_2 = Button_1(self.form2, "Предыдущяя запись", 50, 1, lambda: self.previous_line(table_arr[0]))
        self.but_3 = Button_1(self.form2, "добавить запись", 50, 1, lambda: self.save_data(table_arr[0], rows))
        self.but_4 = Button_1(self.form2, "Удалить запись", 50, 1,
                              lambda: self.del_data(table_arr[0], name_column[0], rows))
        self.but_5 = Button_1(self.form2, "сохранить в файл", 50, 1, lambda: self.extract_text(rows))
        self.but_6 = Button_1(self.form2, "сохранить изменения", 50, 1,
                              lambda: self.save_changes_data(table_arr[0], name_column, rows))
        self.but_7 = Button_1(self.form2, "сохранить в таблицу", 50, 1, lambda: self.save_table(rows, table_arr[0]))
        self.but_8 = Button_1(self.form2, "загрузить из таблицу", 50, 1, lambda: self.load_table(rows, table_arr[0]))
    def table_menu_5(self,table_arr):
        global rows, n
        n = 0
        quest = "SELECT column_name FROM information_schema.columns WHERE table_name = '{0}' order by ordinal_position;".format(table_arr[0])
        name_column = self.quest_def(quest)
        quest = "select * from {0}".format(table_arr[0])
        rows = self.quest_def(quest)
        self.form2 = Form('таблица {0}'.format(table_arr[0]),'500x600+400+200')
        self.qstn_1 = Entry_1(self.form2, '{0}'.format(name_column[0]), rows[0][0], False)
        if len(name_column) > 1:
            self.qstn_2 = Entry_1(self.form2, '{0}'.format(name_column[1]), rows[0][1], False)
            if len(name_column) > 2:
                self.qstn_3 = Entry_1(self.form2, '{0}'.format(name_column[2]), rows[0][2], False)
                if len(name_column) > 3:
                    self.qstn_4 = Entry_1(self.form2, '{0}'.format(name_column[3]), rows[0][3], False)
                    if len(name_column) > 4:
                        self.qstn_5 = Entry_1(self.form2, '{0}'.format(name_column[4]), rows[0][4], False)
                        if len(name_column) > 5:
                            self.qstn_6 = Entry_1(self.form2, '{0}'.format(name_column[5]), rows[0][5], False)
                            if len(name_column) > 6:
                                self.qstn_7 = Entry_1(self.form2, '{0}'.format(name_column[6]), rows[0][6], False)
                                if len(name_column) > 7:
                                    self.qstn_8 = Entry_1(self.form2, '{0}'.format(name_column[7]), rows[0][7], False)
                                    if len(name_column) > 8:
                                        self.qstn_9 = Entry_1(self.form2, '{0}'.format(name_column[8]), rows[0][8], False)
                                        if len(name_column) > 9:
                                            self.qstn_10 = Entry_1(self.form2, '{0}'.format(name_column[9]), rows[0][9], False)
                                            if len(name_column) > 10:
                                                self.qstn_11 = Entry_1(self.form2, '{0}'.format(name_column[10]), rows[0][10], False)
        self.but_1 = Button_1(self.form2, "Следующая запись", 50, 1, lambda: self.next_line(table_arr[0]))
        self.but_2 = Button_1(self.form2, "Предыдущяя запись", 50, 1, lambda: self.previous_line(table_arr[0]))
        self.but_3 = Button_1(self.form2, "добавить запись", 50, 1, lambda: self.save_data(table_arr[0], rows))
        self.but_4 = Button_1(self.form2, "Удалить запись", 50, 1,
                              lambda: self.del_data(table_arr[0], name_column[0], rows))
        self.but_5 = Button_1(self.form2, "сохранить в файл", 50, 1, lambda: self.extract_text(rows))
        self.but_6 = Button_1(self.form2, "сохранить изменения", 50, 1,
                              lambda: self.save_changes_data(table_arr[0], name_column, rows))
        self.but_7 = Button_1(self.form2, "сохранить в таблицу", 50, 1, lambda: self.save_table(rows, table_arr[0]))
        self.but_8 = Button_1(self.form2, "загрузить из таблицу", 50, 1, lambda: self.load_table(rows, table_arr[0]))
    def table_menu_6(self,table_arr):
        global rows, n
        n = 0
        quest = "SELECT column_name FROM information_schema.columns WHERE table_name = '{0}' order by ordinal_position;".format(table_arr[0])
        name_column = self.quest_def(quest)
        quest = "select * from {0}".format(table_arr[0])
        rows = self.quest_def(quest)
        self.form2 = Form('таблица {0}'.format(table_arr[0]),'500x600+400+200')
        self.qstn_1 = Entry_1(self.form2, '{0}'.format(name_column[0]), rows[0][0], False)
        if len(name_column) > 1:
            self.qstn_2 = Entry_1(self.form2, '{0}'.format(name_column[1]), rows[0][1], False)
            if len(name_column) > 2:
                self.qstn_3 = Entry_1(self.form2, '{0}'.format(name_column[2]), rows[0][2], False)
                if len(name_column) > 3:
                    self.qstn_4 = Entry_1(self.form2, '{0}'.format(name_column[3]), rows[0][3], False)
                    if len(name_column) > 4:
                        self.qstn_5 = Entry_1(self.form2, '{0}'.format(name_column[4]), rows[0][4], False)
                        if len(name_column) > 5:
                            self.qstn_6 = Entry_1(self.form2, '{0}'.format(name_column[5]), rows[0][5], False)
                            if len(name_column) > 6:
                                self.qstn_7 = Entry_1(self.form2, '{0}'.format(name_column[6]), rows[0][6], False)
                                if len(name_column) > 7:
                                    self.qstn_8 = Entry_1(self.form2, '{0}'.format(name_column[7]), rows[0][7], False)
                                    if len(name_column) > 8:
                                        self.qstn_9 = Entry_1(self.form2, '{0}'.format(name_column[8]), rows[0][8], False)
                                        if len(name_column) > 9:
                                            self.qstn_10 = Entry_1(self.form2, '{0}'.format(name_column[9]), rows[0][9], False)
                                            if len(name_column) > 10:
                                                self.qstn_11 = Entry_1(self.form2, '{0}'.format(name_column[10]), rows[0][10], False)
        self.but_1 = Button_1(self.form2, "Следующая запись", 50, 1, lambda: self.next_line(table_arr[0]))
        self.but_2 = Button_1(self.form2, "Предыдущяя запись", 50, 1, lambda: self.previous_line(table_arr[0]))
        self.but_3 = Button_1(self.form2, "добавить запись", 50, 1, lambda: self.save_data(table_arr[0], rows))
        self.but_4 = Button_1(self.form2, "Удалить запись", 50, 1,
                              lambda: self.del_data(table_arr[0], name_column[0], rows))
        self.but_5 = Button_1(self.form2, "сохранить в файл", 50, 1, lambda: self.extract_text(rows))
        self.but_6 = Button_1(self.form2, "сохранить изменения", 50, 1,
                              lambda: self.save_changes_data(table_arr[0], name_column, rows))
        self.but_7 = Button_1(self.form2, "сохранить в таблицу", 50, 1, lambda: self.save_table(rows, table_arr[0]))
        self.but_8 = Button_1(self.form2, "загрузить из таблицу", 50, 1, lambda: self.load_table(rows, table_arr[0]))
    def table_menu_7(self,table_arr):
        global rows, n
        n = 0
        quest = "SELECT column_name FROM information_schema.columns WHERE table_name = '{0}' order by ordinal_position;".format(table_arr[0])
        name_column = self.quest_def(quest)
        quest = "select * from {0}".format(table_arr[0])
        rows = self.quest_def(quest)
        self.form2 = Form('таблица {0}'.format(table_arr[0]),'500x600+400+200')
        self.qstn_1 = Entry_1(self.form2, '{0}'.format(name_column[0]), rows[0][0], False)
        if len(name_column) > 1:
            self.qstn_2 = Entry_1(self.form2, '{0}'.format(name_column[1]), rows[0][1], False)
            if len(name_column) > 2:
                self.qstn_3 = Entry_1(self.form2, '{0}'.format(name_column[2]), rows[0][2], False)
                if len(name_column) > 3:
                    self.qstn_4 = Entry_1(self.form2, '{0}'.format(name_column[3]), rows[0][3], False)
                    if len(name_column) > 4:
                        self.qstn_5 = Entry_1(self.form2, '{0}'.format(name_column[4]), rows[0][4], False)
                        if len(name_column) > 5:
                            self.qstn_6 = Entry_1(self.form2, '{0}'.format(name_column[5]), rows[0][5], False)
                            if len(name_column) > 6:
                                self.qstn_7 = Entry_1(self.form2, '{0}'.format(name_column[6]), rows[0][6], False)
                                if len(name_column) > 7:
                                    self.qstn_8 = Entry_1(self.form2, '{0}'.format(name_column[7]), rows[0][7], False)
                                    if len(name_column) > 8:
                                        self.qstn_9 = Entry_1(self.form2, '{0}'.format(name_column[8]), rows[0][8], False)
                                        if len(name_column) > 9:
                                            self.qstn_10 = Entry_1(self.form2, '{0}'.format(name_column[9]), rows[0][9], False)
                                            if len(name_column) > 10:
                                                self.qstn_11 = Entry_1(self.form2, '{0}'.format(name_column[10]), rows[0][10], False)
        self.but_1 = Button_1(self.form2, "Следующая запись", 50, 1, lambda: self.next_line(table_arr[0]))
        self.but_2 = Button_1(self.form2, "Предыдущяя запись", 50, 1, lambda: self.previous_line(table_arr[0]))
        self.but_3 = Button_1(self.form2, "добавить запись", 50, 1, lambda: self.save_data(table_arr[0], rows, name_column[0]))
        self.but_4 = Button_1(self.form2, "Удалить запись", 50, 1,
                              lambda: self.del_data(table_arr[0], name_column[0], rows))
        self.but_5 = Button_1(self.form2, "сохранить в файл", 50, 1, lambda: self.extract_text(rows))
        self.but_6 = Button_1(self.form2, "сохранить изменения", 50, 1,
                              lambda: self.save_changes_data(table_arr[0], name_column, rows))
        self.but_7 = Button_1(self.form2, "сохранить в таблицу", 50, 1, lambda: self.save_table(rows, table_arr[0]))
        self.but_8 = Button_1(self.form2, "загрузить из таблицу", 50, 1, lambda: self.load_table(rows, table_arr[0]))
    def table_menu_8(self,table_arr):
        global rows, n
        n = 0
        quest = "SELECT column_name FROM information_schema.columns WHERE table_name = '{0}' order by ordinal_position;".format(table_arr[0])
        name_column = self.quest_def(quest)
        quest = "select * from {0}".format(table_arr[0])
        rows = self.quest_def(quest)
        self.form2 = Form('таблица {0}'.format(table_arr[0]),'500x600+400+200')
        self.qstn_1 = Entry_1(self.form2, '{0}'.format(name_column[0]), rows[0][0], False)
        if len(name_column) > 1:
            self.qstn_2 = Entry_1(self.form2, '{0}'.format(name_column[1]), rows[0][1], False)
            if len(name_column) > 2:
                self.qstn_3 = Entry_1(self.form2, '{0}'.format(name_column[2]), rows[0][2], False)
                if len(name_column) > 3:
                    self.qstn_4 = Entry_1(self.form2, '{0}'.format(name_column[3]), rows[0][3], False)
                    if len(name_column) > 4:
                        self.qstn_5 = Entry_1(self.form2, '{0}'.format(name_column[4]), rows[0][4], False)
                        if len(name_column) > 5:
                            self.qstn_6 = Entry_1(self.form2, '{0}'.format(name_column[5]), rows[0][5], False)
                            if len(name_column) > 6:
                                self.qstn_7 = Entry_1(self.form2, '{0}'.format(name_column[6]), rows[0][6], False)
                                if len(name_column) > 7:
                                    self.qstn_8 = Entry_1(self.form2, '{0}'.format(name_column[7]), rows[0][7], False)
                                    if len(name_column) > 8:
                                        self.qstn_9 = Entry_1(self.form2, '{0}'.format(name_column[8]), rows[0][8], False)
                                        if len(name_column) > 9:
                                            self.qstn_10 = Entry_1(self.form2, '{0}'.format(name_column[9]), rows[0][9], False)
                                            if len(name_column) > 10:
                                                self.qstn_11 = Entry_1(self.form2, '{0}'.format(name_column[10]), rows[0][10], False)
        self.but_1 = Button_1(self.form2, "Следующая запись", 50, 1, lambda: self.next_line(table_arr[0]))
        self.but_2 = Button_1(self.form2, "Предыдущяя запись", 50, 1, lambda: self.previous_line(table_arr[0]))
        self.but_3 = Button_1(self.form2, "добавить запись", 50, 1, lambda: self.save_data(table_arr[0], rows, name_column[0]))
        self.but_4 = Button_1(self.form2, "Удалить запись", 50, 1,
                              lambda: self.del_data(table_arr[0], name_column[0], rows))
        self.but_5 = Button_1(self.form2, "сохранить в файл", 50, 1, lambda: self.extract_text(rows))
        self.but_6 = Button_1(self.form2, "сохранить изменения", 50, 1,
                              lambda: self.save_changes_data(table_arr[0], name_column, rows))
        self.but_7 = Button_1(self.form2, "сохранить в таблицу", 50, 1, lambda: self.save_table(rows, table_arr[0]))
        self.but_8 = Button_1(self.form2, "загрузить из таблицу", 50, 1, lambda: self.load_table(rows, table_arr[0]))

    def table_menu_9(self,table_arr):
        global rows, n
        n = 0
        quest = "SELECT column_name FROM information_schema.columns WHERE table_name = '{0}';".format(table_arr[0])
        name_column = self.quest_def(quest)
        quest = "select * from {0}".format(table_arr[0])
        rows = self.quest_def(quest)
        self.form2 = Form('таблица {0}'.format(table_arr[0]),'500x600+400+200')
        self.qstn_1 = Entry_1(self.form2, '{0}'.format(name_column[0]), rows[0][0], False)
        if len(name_column) > 1:
            self.qstn_2 = Entry_1(self.form2, '{0}'.format(name_column[1]), rows[0][1], False)
            if len(name_column) > 2:
                self.qstn_3 = Entry_1(self.form2, '{0}'.format(name_column[2]), rows[0][2], False)
                if len(name_column) > 3:
                    self.qstn_4 = Entry_1(self.form2, '{0}'.format(name_column[3]), rows[0][3], False)
                    if len(name_column) > 4:
                        self.qstn_5 = Entry_1(self.form2, '{0}'.format(name_column[4]), rows[0][4], False)
                        if len(name_column) > 5:
                            self.qstn_6 = Entry_1(self.form2, '{0}'.format(name_column[5]), rows[0][5], False)
                            if len(name_column) > 6:
                                self.qstn_7 = Entry_1(self.form2, '{0}'.format(name_column[6]), rows[0][6], False)
                                if len(name_column) > 7:
                                    self.qstn_8 = Entry_1(self.form2, '{0}'.format(name_column[7]), rows[0][7], False)
                                    if len(name_column) > 8:
                                        self.qstn_9 = Entry_1(self.form2, '{0}'.format(name_column[8]), rows[0][8], False)
                                        if len(name_column) > 9:
                                            self.qstn_10 = Entry_1(self.form2, '{0}'.format(name_column[9]), rows[0][9], False)
                                            if len(name_column) > 10:
                                                self.qstn_11 = Entry_1(self.form2, '{0}'.format(name_column[10]), rows[0][10], False)
        self.but_1 = Button_1(self.form2, "Следующая запись", 50, 1, lambda: self.next_line(table_arr[0]))
        self.but_2 = Button_1(self.form2, "Предыдущяя запись", 50, 1, lambda: self.previous_line(table_arr[0]))
        self.but_3 = Button_1(self.form2, "добавить запись", 50, 1, lambda: self.save_data(table_arr[0], rows, name_column))
        self.but_4 = Button_1(self.form2, "Удалить запись", 50, 1,
                              lambda: self.del_data(table_arr[0], name_column[0], rows))
        self.but_5 = Button_1(self.form2, "сохранить в файл", 50, 1, lambda: self.extract_text(rows))
        self.but_6 = Button_1(self.form2, "сохранить изменения", 50, 1,
                              lambda: self.save_changes_data(table_arr[0], name_column, rows))
        self.but_7 = Button_1(self.form2, "сохранить в таблицу", 50, 1, lambda: self.save_table(rows, table_arr[0]))
        self.but_8 = Button_1(self.form2, "загрузить из таблицу", 50, 1, lambda: self.load_table(rows, table_arr[0]))
    def table_menu_10(self,table_arr):
        global rows, n
        n = 0
        quest = "SELECT column_name FROM information_schema.columns WHERE table_name = '{0}';".format(table_arr[0])
        name_column = self.quest_def(quest)
        quest = "select * from {0}".format(table_arr[0])
        rows = self.quest_def(quest)
        self.form2 = Form('таблица {0}'.format(table_arr[0]),'500x600+400+200')
        self.qstn_1 = Entry_1(self.form2, '{0}'.format(name_column[0]), rows[0][0], False)
        if len(name_column) > 1:
            self.qstn_2 = Entry_1(self.form2, '{0}'.format(name_column[1]), rows[0][1], False)
            if len(name_column) > 2:
                self.qstn_3 = Entry_1(self.form2, '{0}'.format(name_column[2]), rows[0][2], False)
                if len(name_column) > 3:
                    self.qstn_4 = Entry_1(self.form2, '{0}'.format(name_column[3]), rows[0][3], False)
                    if len(name_column) > 4:
                        self.qstn_5 = Entry_1(self.form2, '{0}'.format(name_column[4]), rows[0][4], False)
                        if len(name_column) > 5:
                            self.qstn_6 = Entry_1(self.form2, '{0}'.format(name_column[5]), rows[0][5], False)
                            if len(name_column) > 6:
                                self.qstn_7 = Entry_1(self.form2, '{0}'.format(name_column[6]), rows[0][6], False)
                                if len(name_column) > 7:
                                    self.qstn_8 = Entry_1(self.form2, '{0}'.format(name_column[7]), rows[0][7], False)
                                    if len(name_column) > 8:
                                        self.qstn_9 = Entry_1(self.form2, '{0}'.format(name_column[8]), rows[0][8], False)
                                        if len(name_column) > 9:
                                            self.qstn_10 = Entry_1(self.form2, '{0}'.format(name_column[9]), rows[0][9], False)
                                            if len(name_column) > 10:
                                                self.qstn_11 = Entry_1(self.form2, '{0}'.format(name_column[10]), rows[0][10], False)
        self.but_1 = Button_1(self.form2, "Следующая запись", 50, 1, lambda: self.next_line(table_arr[0]))
        self.but_2 = Button_1(self.form2, "Предыдущяя запись", 50, 1, lambda: self.previous_line(table_arr[0]))
        self.but_3 = Button_1(self.form2, "добавить запись", 50, 1, lambda: self.save_data(table_arr[0], rows))
        self.but_4 = Button_1(self.form2, "Удалить запись", 50, 1,
                              lambda: self.del_data(table_arr[0], name_column[0], rows))
        self.but_5 = Button_1(self.form2, "сохранить в файл", 50, 1, lambda: self.extract_text(rows))
        self.but_6 = Button_1(self.form2, "сохранить изменения", 50, 1,
                              lambda: self.save_changes_data(table_arr[0], name_column, rows))
        self.but_7 = Button_1(self.form2, "сохранить в таблицу", 50, 1, lambda: self.save_table(rows, table_arr[0]))
        self.but_8 = Button_1(self.form2, "загрузить из таблицу", 50, 1, lambda: self.load_table(rows, table_arr[0]))
    def table_menu_11(self,table_arr):
        global rows, n
        n = 0
        quest = "SELECT column_name FROM information_schema.columns WHERE table_name = '{0}';".format(table_arr[0])
        name_column = self.quest_def(quest)
        quest = "select * from {0}".format(table_arr[0])
        rows = self.quest_def(quest)
        self.form2 = Form('таблица {0}'.format(table_arr[0]),'500x600+400+200')
        self.qstn_1 = Entry_1(self.form2, '{0}'.format(name_column[0]), rows[0][0], False)
        if len(name_column) > 1:
            self.qstn_2 = Entry_1(self.form2, '{0}'.format(name_column[1]), rows[0][1], False)
            if len(name_column) > 2:
                self.qstn_3 = Entry_1(self.form2, '{0}'.format(name_column[2]), rows[0][2], False)
                if len(name_column) > 3:
                    self.qstn_4 = Entry_1(self.form2, '{0}'.format(name_column[3]), rows[0][3], False)
                    if len(name_column) > 4:
                        self.qstn_5 = Entry_1(self.form2, '{0}'.format(name_column[4]), rows[0][4], False)
                        if len(name_column) > 5:
                            self.qstn_6 = Entry_1(self.form2, '{0}'.format(name_column[5]), rows[0][5], False)
                            if len(name_column) > 6:
                                self.qstn_7 = Entry_1(self.form2, '{0}'.format(name_column[6]), rows[0][6], False)
                                if len(name_column) > 7:
                                    self.qstn_8 = Entry_1(self.form2, '{0}'.format(name_column[7]), rows[0][7], False)
                                    if len(name_column) > 8:
                                        self.qstn_9 = Entry_1(self.form2, '{0}'.format(name_column[8]), rows[0][8], False)
                                        if len(name_column) > 9:
                                            self.qstn_10 = Entry_1(self.form2, '{0}'.format(name_column[9]), rows[0][9], False)
                                            if len(name_column) > 10:
                                                self.qstn_11 = Entry_1(self.form2, '{0}'.format(name_column[10]), rows[0][10], False)
        self.but_1 = Button_1(self.form2, "Следующая запись", 50, 1, lambda: self.next_line(table_arr[0]))
        self.but_2 = Button_1(self.form2, "Предыдущяя запись", 50, 1, lambda: self.previous_line(table_arr[0]))
        self.but_3 = Button_1(self.form2, "добавить запись", 50, 1, lambda: self.save_data(table_arr[0], rows))
        self.but_4 = Button_1(self.form2, "Удалить запись", 50, 1,
                              lambda: self.del_data(table_arr[0], name_column[0], rows))
        self.but_5 = Button_1(self.form2, "сохранить в файл", 50, 1, lambda: self.extract_text(rows))
        self.but_6 = Button_1(self.form2, "сохранить изменения", 50, 1,
                              lambda: self.save_changes_data(table_arr[0], name_column, rows))
        self.but_7 = Button_1(self.form2, "сохранить в таблицу", 50, 1, lambda: self.save_table(rows, table_arr[0]))
        self.but_8 = Button_1(self.form2, "загрузить из таблицу", 50, 1, lambda: self.load_table(rows, table_arr[0]))

    def win_close():
        global n
        n = 0
        print("sadg,")
    def next_line(self, table_name):
        global rows
        global n
        if len(rows) - 1 > n:
            n = n + 1
        else: self.show_error("это последняя строка")
        if len(rows[0]) > 0:
            self.qstn_1.set_1(rows[n][0])
            if len(rows[0]) > 1:
                self.qstn_2.set_1(rows[n][1])
                if len(rows[0]) > 2:
                    self.qstn_3.set_1(rows[n][2])
                    if len(rows[0]) > 3:
                        self.qstn_4.set_1(rows[n][3])
                        if len(rows[0]) > 4:
                            self.qstn_5.set_1(rows[n][4])
                            if len(rows[0]) > 5:
                                self.qstn_6.set_1(rows[n][5])
                                if len(rows[0]) > 6:
                                    self.qstn_7.set_1(rows[n][6])
                                    if len(rows[0]) > 7:
                                        self.qstn_8.set_1(rows[n][7])
                                        if len(rows[0]) > 8:
                                            self.qstn_9.set_1(rows[n][8])
                                            if len(rows[0]) > 9:
                                                self.qstn_10.set_1(rows[n][9])
                                                if len(rows[0]) > 10:
                                                    self.qstn_11.set_1(rows[n][10])
                                                    if len(rows[0]) > 11:
                                                        self.qstn_12.set_1(rows[n][11])
    def previous_line(self, table_name):
        global n, rows
        if n > 0:
            n = n - 1
            if len(rows[0]) > 0:
                self.qstn_1.set_1(rows[n][0])
                if len(rows[0]) > 1:
                    self.qstn_2.set_1(rows[n][1])
                    if len(rows[0]) > 2:
                        self.qstn_3.set_1(rows[n][2])
                        if len(rows[0]) > 3:
                            self.qstn_4.set_1(rows[n][3])
                            if len(rows[0]) > 4:
                                self.qstn_5.set_1(rows[n][4])
                                if len(rows[0]) > 5:
                                    self.qstn_6.set_1(rows[n][5])
                                    if len(rows[0]) > 6:
                                        self.qstn_7.set_1(rows[n][6])
                                        if len(rows[0]) > 7:
                                            self.qstn_8.set_1(rows[n][7])
                                            if len(rows[0]) > 8:
                                                self.qstn_9.set_1(rows[n][8])
                                                if len(rows[0]) > 9:
                                                    self.qstn_10.set_1(rows[n][9])
                                                    if len(rows[0]) > 10:
                                                        self.qstn_11.set_1(rows[n][10])
                                                        if len(rows[0]) > 11:
                                                            self.qstn_12.set_1(rows[n][11])
        else:
            self.show_error("это первая строка")
        self.connect(table_name)
    def save_data(self, table_name, rows, column_name):
        global n
        temp = []
        quest = "select * from {0} where {1} = {2}".format(table_name, str(column_name)[2:-3], self.qstn_1.get_1())
        row = self.quest_def(quest)
        if str(row) != "[]":
            self.show_error("строка с таким id уже существует")
        else:
            if len(rows[0]) > 0:
                x = self.qstn_1.get_1()
                temp.append(x)
                if len(rows[0]) > 1:
                    x = self.qstn_2.get_1()
                    temp.append(x)
                    if len(rows[0]) > 2:
                        x = self.qstn_3.get_1()
                        temp.append(x)
                        if len(rows[0]) > 3:
                            x = self.qstn_4.get_1()
                            temp.append(x)
                            if len(rows[0]) > 4:
                                x = self.qstn_5.get_1()
                                temp.append(x)
                                if len(rows[0]) > 5:
                                    x = self.qstn_6.get_1()
                                    temp.append(x)
                                    if len(rows[0]) > 6:
                                        x = self.qstn_7.get_1()
                                        temp.append(x)
                                        if len(rows[0]) > 7:
                                            x = self.qstn_8.get_1()
                                            temp.append(x)
                                            if len(rows[0]) > 8:
                                                x = self.qstn_9.get_1()
                                                temp.append(x)
                                                if len(rows[0]) > 9:
                                                    x = self.qstn_10.get_1()
                                                    temp.append(x)
                                                    if len(rows[0]) > 10:
                                                        x = self.qstn_11.get_1()
                                                        temp.append(x)
                                                        if len(rows[0]) > 11:
                                                            x = self.qstn_12.get_1()
                                                            temp.append(x)
            n += 1
        global my_database, my_user, my_password, my_host, my_port
        conn = psycopg2.connect(database=my_database, user=my_user,password=my_password, host=my_host,port=my_port)
        cur = conn.cursor()
        if len(rows[0]) > 0:
            quest = "INSERT INTO {0} VALUES ('{1}')".format(table_name,temp[0])
            if len(rows[0]) > 1:
                quest = "INSERT INTO {0} VALUES ('{1}','{2}')".format(table_name, temp[0],temp[1])
                if len(rows[0]) > 2:
                    quest = "INSERT INTO {0} VALUES ('{1}','{2}','{3}')".format(table_name, temp[0], temp[1], temp[2])
                    if len(rows[0]) > 3:
                        quest = "INSERT INTO {0} VALUES ('{1}','{2}','{3}','{4}')".format(table_name, temp[0], temp[1], temp[2], temp[3])
                        if len(rows[0]) > 4:
                            quest = "INSERT INTO {0} VALUES ('{1}','{2}','{3}','{4}','{5}')".format(table_name, temp[0], temp[1], temp[2], temp[3], temp[4])
                            if len(rows[0]) > 5:
                                quest = "INSERT INTO {0} VALUES ('{1}','{2}','{3}','{4}','{5}','{6}')".format(table_name, temp[0], temp[1], temp[2], temp[3], temp[4], temp[5])
                                if len(rows[0]) > 6:
                                    quest = "INSERT INTO {0} VALUES ('{1}','{2}','{3}','{4}','{5}','{6}','{7}')".format(table_name, temp[0], temp[1], temp[2], temp[3], temp[4], temp[5],temp[6])
                                    if len(rows[0]) > 7:
                                        quest = "INSERT INTO {0} VALUES ('{1}','{2}','{3}','{4}','{5}','{6}','{7}','{8}')".format(table_name, temp[0], temp[1], temp[2], temp[3], temp[4], temp[5],temp[6], temp[7])
                                        if len(rows[0]) > 8:
                                            quest = "INSERT INTO {0} VALUES ('{1}','{2}','{3}','{4}','{5}','{6}','{7}','{8}','{9}')".format(table_name, temp[0], temp[1], temp[2], temp[3], temp[4], temp[5],temp[6], temp[7], temp[8])
                                            if len(rows[0]) > 9:
                                                quest = "INSERT INTO {0} VALUES ('{1}','{2}','{3}','{4}','{5}','{6}','{7}','{8}','{9}','{10}')".format(table_name,temp[0], temp[1], temp[2], temp[3], temp[4], temp[5],temp[6], temp[7], temp[8],temp[9])
                                                if len(rows[0]) > 10:
                                                    quest = "INSERT INTO {0} VALUES ('{1}','{2}','{3}','{4}','{5}','{6}','{7}','{8}','{9}','{10}','{11}')".format(table_name, temp[0], temp[1], temp[2], temp[3], temp[4], temp[5],temp[6], temp[7], temp[8], temp[9], temp[10])
                                                    if len(rows[0]) > 11:
                                                        quest = "INSERT INTO {0} VALUES ('{1}','{2}','{3}','{4}','{5}','{6}','{7}','{8}','{9}','{10}','{11}','{12}')".format(table_name, temp[0], temp[1], temp[2], temp[3], temp[4], temp[5],temp[6], temp[7], temp[8], temp[9], temp[10], temp[11])
        cur.execute(quest)
        conn.commit()
        conn.close()
    def del_data(self, table_name, name_column):
        global my_database, my_user, my_password, my_host, my_port
        conn = psycopg2.connect(database=my_database, user=my_user,password=my_password, host=my_host,port=my_port)
        cur = conn.cursor()
        x = self.qstn_1.get_1()
        quest = "DELETE FROM {0} WHERE {1} = '{2}' ".format(table_name, str(name_column)[2:-3], x)
        cur.execute(quest)
        conn.commit()
        conn.close()
        self.previous_line(table_name)
    def save_changes_data(self, table_name, name_column, rows):
        temp = []
        if len(rows[0]) > 0:
            x = self.qstn_1.get_1()
            temp.append(x)
            if len(rows[0]) > 1:
                x = self.qstn_2.get_1()
                temp.append(x)
                if len(rows[0]) > 2:
                    x = self.qstn_3.get_1()
                    temp.append(x)
                    if len(rows[0]) > 3:
                        x = self.qstn_4.get_1()
                        temp.append(x)
                        if len(rows[0]) > 4:
                            x = self.qstn_5.get_1()
                            temp.append(x)
                            if len(rows[0]) > 5:
                                x = self.qstn_6.get_1()
                                temp.append(x)
                                if len(rows[0]) > 6:
                                    x = self.qstn_7.get_1()
                                    temp.append(x)
                                    if len(rows[0]) > 7:
                                        x = self.qstn_8.get_1()
                                        temp.append(x)
                                        if len(rows[0]) > 8:
                                            x = self.qstn_9.get_1()
                                            temp.append(x)
                                            if len(rows[0]) > 9:
                                                x = self.qstn_10.get_1()
                                                temp.append(x)
                                                if len(rows[0]) > 10:
                                                    x = self.qstn_11.get_1()
                                                    temp.append(x)
                                                    if len(rows[0]) > 11:
                                                        x = self.qstn_12.get_1()
                                                        temp.append(x)
        global my_database, my_user, my_password, my_host, my_port
        conn = psycopg2.connect(database=my_database, user=my_user, password=my_password, host=my_host, port=my_port)
        cur = conn.cursor()
        if len(temp) > 1:
            quest = "UPDATE {0} SET {3} = '{4}' where {1} = '{2}'".format(table_name, str(name_column[0])[2:-3], temp[0], str(name_column[1])[2:-3], temp[1] )
            if len(temp) > 2:
                quest = "UPDATE {0} SET {3} = '{4}', {5} = '{6}' where {1} = '{2}'".format(table_name, str(name_column[0])[2:-3], temp[0], str(name_column[1])[2:-3], temp[1], str(name_column[2])[2:-3], temp[2])
                if len(temp) > 3:
                    quest = "UPDATE {0} SET {3} = '{4}', {5} = '{6}', {7} = {8} where {1} = '{2}'".format(table_name, str(name_column[0])[2:-3], temp[0], str(name_column[1])[2:-3], temp[1], str(name_column[2])[2:-3], temp[2], str(name_column[3])[2:-3], temp[3])
                    if len(temp) > 4:
                        quest = "UPDATE {0} SET {3} = '{4}', {5} = '{6}', {7} = '{8}', {9} = '{10}' where {1} = '{2}'".format(table_name, str(name_column[0])[2:-3], temp[0], str(name_column[1])[2:-3], temp[1], str(name_column[2])[2:-3], temp[2], str(name_column[3])[2:-3], temp[3], str(name_column[4])[2:-3], temp[4])
                        if len(temp) > 5:
                            quest = "UPDATE {0} SET {3} = '{4}', {5} = '{6}', {7} = '{8}', {9} = '{10}', {11} = '{12}' where {1} = '{2}'".format(table_name, str(name_column[0])[2:-3], temp[0], str(name_column[1])[2:-3], temp[1],
                                str(name_column[2])[2:-3], temp[2], str(name_column[3])[2:-3], temp[3],
                                str(name_column[4])[2:-3], temp[4],str(name_column[5])[2:-3], temp[5])
                            if len(temp) > 6:
                                quest = "UPDATE {0} SET {3} = '{4}', {5} = '{6}', {7} = '{8}', {9} = '{10}', {11} = '{12}', {13} = '{14}' where {1} = '{2}'".format(
                                    table_name, str(name_column[0])[2:-3], temp[0], str(name_column[1])[2:-3], temp[1],
                                    str(name_column[2])[2:-3], temp[2], str(name_column[3])[2:-3], temp[3],
                                    str(name_column[4])[2:-3], temp[4], str(name_column[5])[2:-3], temp[5], str(name_column[6])[2:-3], temp[6])
                                if len(temp) > 7:
                                    quest = "UPDATE {0} SET {3} = '{4}', {5} = '{6}', {7} = '{8}', {9} = '{10}', {11} = '{12}', {13} = '{14}', {15} = '{16}' where {1} = '{2}'".format(
                                        table_name, str(name_column[0])[2:-3], temp[0], str(name_column[1])[2:-3],
                                        temp[1],
                                        str(name_column[2])[2:-3], temp[2], str(name_column[3])[2:-3], temp[3],
                                        str(name_column[4])[2:-3], temp[4], str(name_column[5])[2:-3], temp[5],
                                        str(name_column[6])[2:-3], temp[6],
                                        str(name_column[7])[2:-3], temp[7])
                                    if len(temp) > 8:
                                        quest = "UPDATE {0} SET {3} = '{4}', {5} = '{6}', {7} = '{8}', {9} = '{10}', {11} = '{12}', {13} = '{14}', {15} = '{16}', {17} = '{18}' where {1} = '{2}'".format(
                                            table_name, str(name_column[0])[2:-3], temp[0], str(name_column[1])[2:-3],
                                            temp[1],
                                            str(name_column[2])[2:-3], temp[2], str(name_column[3])[2:-3], temp[3],
                                            str(name_column[4])[2:-3], temp[4], str(name_column[5])[2:-3], temp[5],
                                            str(name_column[6])[2:-3], temp[6],
                                            str(name_column[7])[2:-3], temp[7],
                                            str(name_column[8])[2:-3], temp[8])
                                        if len(temp) > 9:
                                            quest = "UPDATE {0} SET {3} = '{4}', {5} = '{6}', {7} = '{8}', {9} = '{10}', {11} = '{12}', {13} = '{14}', {15} = '{16}', {17} = '{18}', {19} = '{20}' where {1} = '{2}'".format(
                                                table_name, str(name_column[0])[2:-3], temp[0],
                                                str(name_column[1])[2:-3],
                                                temp[1],
                                                str(name_column[2])[2:-3], temp[2], str(name_column[3])[2:-3], temp[3],
                                                str(name_column[4])[2:-3], temp[4], str(name_column[5])[2:-3], temp[5],
                                                str(name_column[6])[2:-3], temp[6],
                                                str(name_column[7])[2:-3], temp[7],
                                                str(name_column[8])[2:-3], temp[8],
                                                str(name_column[9])[2:-3], temp[9])
                                            if len(temp) > 10:
                                                quest = "UPDATE {0} SET {3} = '{4}', {5} = '{6}', {7} = '{8}', {9} = '{10}', {11} = '{12}', {13} = '{14}', {15} = '{16}', {17} = '{18}', {19} = '{20}', {21} = '{22}' where {1} = '{2}'".format(
                                                    table_name, str(name_column[0])[2:-3], temp[0],
                                                    str(name_column[1])[2:-3],
                                                    temp[1],
                                                    str(name_column[2])[2:-3], temp[2], str(name_column[3])[2:-3],
                                                    temp[3],
                                                    str(name_column[4])[2:-3], temp[4], str(name_column[5])[2:-3],
                                                    temp[5],
                                                    str(name_column[6])[2:-3], temp[6],
                                                    str(name_column[7])[2:-3], temp[7],
                                                    str(name_column[8])[2:-3], temp[8],
                                                    str(name_column[9])[2:-3], temp[9],
                                                    str(name_column[10])[2:-3], temp[10])
                                                if len(temp) > 11:
                                                    quest = "UPDATE {0} SET {3} = '{4}', {5} = '{6}', {7} = '{8}', {9} = '{10}', {11} = '{12}', {13} = '{14}', {15} = '{16}', {17} = '{18}', {19} = '{20}', {21} = '{22}', {23} = '{24}' where {1} = '{2}'".format(
                                                        table_name, str(name_column[0])[2:-3], temp[0],
                                                        str(name_column[1])[2:-3],
                                                        temp[1],
                                                        str(name_column[2])[2:-3], temp[2], str(name_column[3])[2:-3],
                                                        temp[3],
                                                        str(name_column[4])[2:-3], temp[4], str(name_column[5])[2:-3],
                                                        temp[5],
                                                        str(name_column[6])[2:-3], temp[6],
                                                        str(name_column[7])[2:-3], temp[7],
                                                        str(name_column[8])[2:-3], temp[8],
                                                        str(name_column[9])[2:-3], temp[9],
                                                        str(name_column[10])[2:-3], temp[10],
                                                        str(name_column[11])[2:-3], temp[11])
        cur.execute(quest)
        conn.commit()
        conn.close()
    def extract_text(self, rows):
        temp = []
        if len(rows[0]) > 0:
            x = self.qstn_1.get_1()
            temp.append(x)
            if len(rows[0]) > 1:
                x = self.qstn_2.get_1()
                temp.append(x)
                if len(rows[0]) > 2:
                    x = self.qstn_3.get_1()
                    temp.append(x)
                    if len(rows[0]) > 3:
                        x = self.qstn_4.get_1()
                        temp.append(x)
                        if len(rows[0]) > 4:
                            x = self.qstn_5.get_1()
                            temp.append(x)
                            if len(rows[0]) > 5:
                                x = self.qstn_6.get_1()
                                temp.append(x)
                                if len(rows[0]) > 6:
                                    x = self.qstn_7.get_1()
                                    temp.append(x)
                                    if len(rows[0]) > 7:
                                        x = self.qstn_8.get_1()
                                        temp.append(x)
                                        if len(rows[0]) > 8:
                                            x = self.qstn_9.get_1()
                                            temp.append(x)
                                            if len(rows[0]) > 9:
                                                x = self.qstn_10.get_1()
                                                temp.append(x)
                                                if len(rows[0]) > 10:
                                                    x = self.qstn_11.get_1()
                                                    temp.append(x)
                                                    if len(rows[0]) > 11:
                                                        x = self.qstn_12.get_1()
                                                        temp.append(x)
        file_name = fd.asksaveasfilename(defaultextension=".*",
                                         filetypes=(("TXT files", "*.txt"),
                                                    ("HTML files", "*.html;*.htm"),
                                                    ("json files", "*.json"),
                                                    ("xml files","*.xml" ),
                                                    ("all files","*.*")))
        file_type = file_name.split('.')[-1]
        if file_type == "txt" or file_type == "htm" or file_type == "html":
            f = open(file_name, 'w')
            f.write(str(temp))
        elif file_type == "json":
            with open(file_name, "w") as files:
                files.write(json.dumps(temp, sort_keys=True, indent=2, ensure_ascii=False))
        elif file_type == "xml":
            xml_txt = dicttoxml(temp, attr_type=False)
            value = xml_txt.decode()
            with open(file_name, "w") as file:
                file.write(value)
    def save_table(self, rows, table_arr):
        file_name = fd.asksaveasfilename(
            filetypes=(("EXCEL files", "*.xlsx"),
                       ("ALL files", "*.*")))
        wb = openpyxl.load_workbook(file_name)
        if table_arr in wb.sheetnames:
            worksheet = wb[table_arr]
        else:
            worksheet = wb.create_sheet(table_arr)
        temp = []
        if len(rows[0]) > 0:
            x = self.qstn_1.get_1()
            temp.append(x)
            if len(rows[0]) > 1:
                x = self.qstn_2.get_1()
                temp.append(x)
                if len(rows[0]) > 2:
                    x = self.qstn_3.get_1()
                    temp.append(x)
                    if len(rows[0]) > 3:
                        x = self.qstn_4.get_1()
                        temp.append(x)
                        if len(rows[0]) > 4:
                            x = self.qstn_5.get_1()
                            temp.append(x)
                            if len(rows[0]) > 5:
                                x = self.qstn_6.get_1()
                                temp.append(x)
                                if len(rows[0]) > 6:
                                    x = self.qstn_7.get_1()
                                    temp.append(x)
                                    if len(rows[0]) > 7:
                                        x = self.qstn_8.get_1()
                                        temp.append(x)
                                        if len(rows[0]) > 8:
                                            x = self.qstn_9.get_1()
                                            temp.append(x)
                                            if len(rows[0]) > 9:
                                                x = self.qstn_10.get_1()
                                                temp.append(x)
                                                if len(rows[0]) > 10:
                                                    x = self.qstn_11.get_1()
                                                    temp.append(x)
                                                    if len(rows[0]) > 11:
                                                        x = self.qstn_12.get_1()
                                                        temp.append(x)
        worksheet.append(temp)
        wb.save(file_name)

    def show_error(self, text):
        mb.showerror("Ошибка", text)

    def open_id(self,table_arr, name_column, rows):
        id = int(self.qstn_1.get_1())
        global my_database, my_user, my_password, my_host, my_port
        quest = "select * from {0} where {1} = {2}".format(str(table_arr)[2:-3], str(name_column)[2:-3], id)
        row = self.quest_def(quest)
        if str(row) == "[]":
            self.show_error("вызвана несуществующая строка")
        if len(rows[0]) > 1:
            self.qstn_1.set_1(row[0][0])
            if len(rows[0]) > 2:
                self.qstn_2.set_1(row[0][1])
                if len(rows[0]) > 3:
                    self.qstn_3.set_1(row[0][2])
                    if len(rows[0]) > 4:
                        self.qstn_4.set_1(row[0][3])
                        if len(rows[0]) > 5:
                            self.qstn_5.set_1(row[0][4])
                            if len(rows[0]) > 6:
                                self.qstn_6.set_1(row[0][5])
                                if len(rows[0]) > 7:
                                    self.qstn_7.set_1(row[0][6])
                                    if len(rows[0]) > 8:
                                        self.qstn_8.set_1(row[0][7])
                                        if len(rows[0]) > 9:
                                            self.qstn_9.set_1(row[0][8])
                                            if len(rows[0]) > 10:
                                                self.qstn_10.set_1(row[0][9])
                                                if len(rows[0]) > 11:
                                                    self.qstn_11.set_1(row[0][10])
                                                    if len(rows[0]) > 12:
                                                        self.qstn_12.set_1(row[0][11])

    def load_table(self, rows, table_arr):
        file_name = fd.asksaveasfilename(
            filetypes=(("EXCEL files", "*.xlsx"),
                       ("ALL files", "*.*")))
        wb = openpyxl.load_workbook(file_name)
        sheet = wb[table_arr]
        temp = []
        for row in sheet.rows:
            for cell in row:
                temp.append(cell.value)
        if len(rows[0]) >= 1:
            self.qstn_1.set_1(temp[0])
            if len(rows[0]) >= 2:
                self.qstn_2.set_1(temp[1])
                if len(rows[0]) >= 3:
                    self.qstn_3.set_1(temp[2])
                    if len(rows[0]) >= 4:
                        self.qstn_4.set_1(temp[3])
                        if len(rows[0]) >= 5:
                            self.qstn_5.set_1(temp[4])
                            if len(rows[0]) >= 6:
                                self.qstn_6.set_1(temp[5])
                                if len(rows[0]) >= 7:
                                    self.qstn_7.set_1(temp[6])
                                    if len(rows[0]) >= 8:
                                        self.qstn_8.set_1(temp[7])
                                        if len(rows[0]) >= 9:
                                            self.qstn_9.set_1(temp[8])
                                            if len(rows[0]) >= 10:
                                                self.qstn_10.set_1(temp[9])
                                                if len(rows[0]) >= 11:
                                                    self.qstn_11.set_1(temp[10])
                                                    if len(rows[0]) >= 12:
                                                        self.qstn_12.set_1(temp[11])
scene = Scene_1()